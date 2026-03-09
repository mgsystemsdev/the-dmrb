from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

TABLE_STYLE = "TableStyleMedium15"
DATE_FORMAT = "MM/DD/YY"
NUMBER_FORMAT = "#,##0"
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLOR_HEX = {
    "green": "E2EFDA",
    "amber": "FFF3E0",
    "red": "FCE4EC",
    "blue": "E3F2FD",
    "gray": "F5F5F5",
    "yellow": "FFF9C4",
    "header_blue": "4472C4",
}

FILL = {
    k: PatternFill(fill_type="solid", start_color=v, end_color=v) for k, v in COLOR_HEX.items()
}


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def status_fill_name(status: str) -> str | None:
    s = (status or "").strip().lower()
    if "vacant not ready" in s or "not ready" in s:
        return "red"
    if "vacant ready" in s or s == "ready":
        return "green"
    if "on notice" in s or "notice" in s:
        return "gray"
    return None


def alert_fill_name(alert: str) -> str | None:
    s = (alert or "").upper()
    if "CRITICAL MOVE-IN RISK" in s or "WORK STALLED" in s:
        return "red"
    if "SCHEDULED TO MOVE IN" in s or "APARTMENT READY" in s:
        return "green"
    if "NEEDS ATTENTION" in s:
        return "amber"
    if "IN PROGRESS" in s:
        return "blue"
    if "ON NOTICE" in s or "CANCELED" in s or "CLOSED" in s:
        return "gray"
    return None


def dv_fill_name(dv_value) -> str | None:
    try:
        dv = int(dv_value)
    except (TypeError, ValueError):
        return None
    if dv <= 5:
        return "green"
    if dv <= 20:
        return "amber"
    return "red"


def progress_fill_name(progress) -> str | None:
    try:
        val = float(progress)
    except (TypeError, ValueError):
        return None
    if val >= 75:
        return "green"
    if val >= 25:
        return "amber"
    return "red"


def task_status_fill_name(status: str) -> str | None:
    s = (status or "").strip().upper()
    if s in {"VENDOR_COMPLETED", "COMPLETED", "NA"}:
        return "green"
    if s in {"IN_PROGRESS", "SCHEDULED", "IN PROGRESS"}:
        return "blue"
    if s in {"NOT_STARTED", "NOT STARTED"}:
        return "red"
    return None


def wd_fill_name(wd_summary: str) -> str | None:
    s = (wd_summary or "").strip().upper()
    if any(token in s for token in ("OK", "INSTALLED", "CONFIRMED")):
        return "green"
    if any(token in s for token in ("NOTIFIED", "ORDERED")):
        return "amber"
    if any(token in s for token in ("PENDING", "MISSING", "NEEDED")):
        return "red"
    return None


def sla_compliance_fill_name(value) -> str | None:
    try:
        pct = float(value)
    except (TypeError, ValueError):
        return None
    if pct >= 90:
        return "green"
    if pct >= 70:
        return "yellow"
    return "red"


class WorkbookWriter:
    def __init__(self, workbook: Workbook | None = None):
        self.wb = workbook or Workbook()
        self.table_counter = 1

    def new_sheet(self, title: str):
        if len(self.wb.worksheets) == 1 and self.wb.active.max_row == 1 and self.wb.active.max_column == 1:
            ws = self.wb.active
            ws.title = title
            return ws
        return self.wb.create_sheet(title=title)

    def write_section_title(self, ws, row: int, title: str, colspan: int) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(colspan, 1))
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=13, color="1F4E79")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        return row + 1

    def write_table(self, ws, row: int, headers: list[str], rows: Iterable[list], *, as_table: bool = True) -> tuple[int, int]:
        start_row = row
        for col_idx, header in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=header)
            c.font = Font(bold=True)
            c.alignment = ALIGN_CENTER_WRAP
        row += 1

        data_start = row
        for data in rows:
            for col_idx, value in enumerate(data, start=1):
                c = ws.cell(row=row, column=col_idx, value=value)
                c.alignment = ALIGN_CENTER_WRAP
                if isinstance(value, (date, datetime)):
                    c.number_format = DATE_FORMAT
                elif isinstance(value, (int, float)):
                    c.number_format = NUMBER_FORMAT
            row += 1

        end_row = max(data_start, row - 1)
        if as_table and end_row >= start_row + 1:
            ref = f"A{start_row}:{get_column_letter(len(headers))}{end_row}"
            table = Table(displayName=f"Tbl{self.table_counter}", ref=ref)
            self.table_counter += 1
            table.tableStyleInfo = TableStyleInfo(
                name=TABLE_STYLE,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
        return row, end_row

    def apply_fill(self, ws, row: int, col: int, fill_name: str | None) -> None:
        if fill_name and fill_name in FILL:
            ws.cell(row=row, column=col).fill = FILL[fill_name]

    def write_empty_table_header(self, ws, row: int, headers: list[str]) -> int:
        for col_idx, header in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=header)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = FILL["header_blue"]
            c.alignment = ALIGN_CENTER_WRAP
        return row + 1

    def write_empty_message(self, ws, row: int, message: str, colspan: int) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(colspan, 1))
        c = ws.cell(row=row, column=1, value=message)
        c.font = Font(italic=True, color="808080")
        c.alignment = Alignment(horizontal="left", vertical="center")
        return row + 2

    def auto_size_columns(self, ws) -> None:
        for col in ws.columns:
            col_idx = col[0].column
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()
