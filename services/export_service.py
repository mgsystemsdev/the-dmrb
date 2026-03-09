from __future__ import annotations

from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from io import BytesIO
from statistics import mean
import zipfile

from openpyxl import Workbook

from db import repository
from services import board_query_service
from services.excel_writer import (
    WorkbookWriter,
    alert_fill_name,
    dv_fill_name,
    parse_date,
    progress_fill_name,
    sla_compliance_fill_name,
    status_fill_name,
    task_status_fill_name,
    wd_fill_name,
    workbook_to_bytes,
)

FINAL_REPORT_FILENAME = "Final_Report.xlsx"
DMRB_REPORT_FILENAME = "DMRB_Report.xlsx"
DASHBOARD_CHART_FILENAME = "Dashboard_Chart.png"
WEEKLY_SUMMARY_FILENAME = "Weekly_Summary.txt"
ALL_REPORTS_FILENAME = "DMRB_Reports.zip"

TASK_COLUMNS = [
    ("Inspection", "task_insp"),
    ("Paint", "task_paint"),
    ("Make Ready", "task_mr"),
    ("Housekeeping", "task_hk"),
    ("Carpet Clean", "task_cc"),
]
TASK_LABEL_TO_KEY = {
    "INSP": "task_insp",
    "INSPECTION": "task_insp",
    "PAINT": "task_paint",
    "MR": "task_mr",
    "MAKE READY": "task_mr",
    "HK": "task_hk",
    "HOUSEKEEPING": "task_hk",
    "CC": "task_cc",
    "CARPET CLEAN": "task_cc",
    "FW": "task_fw",
    "FINAL WALK": "task_fw",
    "QC": "task_qc",
}


def _safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _has_move_in(row: dict) -> bool:
    return parse_date(row.get("move_in_date")) is not None


def _phase_token(row: dict) -> str:
    return str(row.get("phase") or row.get("nvm") or "").upper()


def _is_vacant(row: dict) -> bool:
    token = _phase_token(row)
    nvm = str(row.get("nvm") or "").upper()
    return token == "VACANT" or nvm == "VACANT"


def _is_notice(row: dict) -> bool:
    token = _phase_token(row)
    nvm = str(row.get("nvm") or "").upper()
    return "NOTICE" in token or "NOTICE" in nvm


def _status_label(row: dict) -> str:
    return str(row.get("manual_ready_status") or "")


def _status_is_ready(row: dict) -> bool:
    s = _status_label(row).lower()
    return "ready" in s and "not ready" not in s


def _status_is_not_ready(row: dict) -> bool:
    s = _status_label(row).lower()
    return "not ready" in s or ("ready" not in s)


def _phase_display(row: dict) -> str:
    return str(row.get("building") or "")


def _unit_display(row: dict) -> str:
    return str(row.get("unit_code") or "")


def _derive_qc_status(row: dict) -> str:
    fw = row.get("task_fw") or {}
    if not fw:
        return "N/A"
    conf = str(fw.get("confirmation_status") or "").upper()
    if conf == "CONFIRMED":
        return "Confirmed"
    if conf == "REJECTED":
        return "Rejected"
    return "Pending"


def _derive_wd_summary(row: dict) -> str:
    wd_present = bool(row.get("wd_present"))
    wd_installed = bool(row.get("wd_installed"))
    wd_notified = bool(row.get("wd_supervisor_notified"))
    if not wd_present:
        return "—"
    if wd_installed:
        return "OK"
    if wd_notified:
        return "NOTIFIED"
    return "PENDING"


def _task_key_from_label(label: str) -> str | None:
    if not label:
        return None
    normalized = str(label).strip().upper()
    return TASK_LABEL_TO_KEY.get(normalized)


def _task_execution_status(row: dict, task_key: str) -> str:
    task = row.get(task_key) or {}
    return str(task.get("execution_status") or "")


def _task_due_date(row: dict, task_key: str):
    task = row.get(task_key) or {}
    return parse_date(task.get("vendor_due_date"))


def _build_notes_lookup(conn, turnover_ids: list[int]) -> dict[int, str]:
    notes = repository.get_notes_for_turnover_ids(conn, turnover_ids, unresolved_only=False)
    notes_by_tid: dict[int, list[str]] = defaultdict(list)
    for note in notes:
        tid = note.get("turnover_id")
        text = (note.get("description") or "").strip()
        if tid is not None and text:
            notes_by_tid[int(tid)].append(text)
    return {tid: "; ".join(texts) for tid, texts in notes_by_tid.items()}


def build_export_turnovers(conn, today: date | None = None) -> list[dict]:
    today = today or date.today()
    rows = board_query_service.get_dmrb_board_rows(conn, today=today)
    turnover_ids = [r.get("turnover_id") for r in rows if r.get("turnover_id") is not None]
    notes_lookup = _build_notes_lookup(conn, turnover_ids)
    out: list[dict] = []
    for row in rows:
        item = dict(row)
        move_in = parse_date(item.get("move_in_date"))
        item["qc_status"] = _derive_qc_status(item)
        item["wd_summary"] = _derive_wd_summary(item)
        item["notes_joined"] = notes_lookup.get(item.get("turnover_id"), item.get("notes_text") or "")
        item["available_date"] = item.get("confirmed_move_out_date") or item.get("move_out_date")
        if item.get("days_to_move_in") is None:
            item["days_to_move_in"] = (move_in - today).days if move_in else None
        out.append(item)
    return out


def _apply_status_fill(writer: WorkbookWriter, ws, row_idx: int, col_idx: int, status: str) -> None:
    writer.apply_fill(ws, row_idx, col_idx, status_fill_name(status))


def _build_final_report(turnovers: list[dict]) -> bytes:
    wb = Workbook()
    writer = WorkbookWriter(wb)

    headers = ["Phase", "Unit", "Status", "Available Date", "Move-In Ready Date", "Move In Date", "MO/Confirm"]

    def build_row(t: dict):
        return [
            _phase_display(t),
            _unit_display(t),
            _status_label(t),
            parse_date(t.get("available_date")),
            parse_date(t.get("report_ready_date")),
            parse_date(t.get("move_in_date")),
            "Yes" if t.get("confirmed_move_out_date") else "",
        ]

    # 1) Reconciliation
    ws = writer.new_sheet("Reconciliation")
    data_rows = [build_row(t) for t in turnovers]
    _, end_row = writer.write_table(ws, 1, headers, data_rows)
    for i, t in enumerate(turnovers, start=2):
        _apply_status_fill(writer, ws, i, 3, _status_label(t))
        if t.get("confirmed_move_out_date"):
            writer.apply_fill(ws, i, 7, "green")
    writer.auto_size_columns(ws)

    # 2) Split View
    ws = writer.new_sheet("Split View")
    row = 1
    with_move_in = [t for t in turnovers if _has_move_in(t)]
    without_move_in = [t for t in turnovers if not _has_move_in(t)]
    row = writer.write_section_title(ws, row, "Has Move In", len(headers))
    row, end_row = writer.write_table(ws, row, headers, [build_row(t) for t in with_move_in])
    for i, t in enumerate(with_move_in, start=row - len(with_move_in)):
        _apply_status_fill(writer, ws, i, 3, _status_label(t))
        if t.get("confirmed_move_out_date"):
            writer.apply_fill(ws, i, 7, "green")
    row = end_row + 2
    row = writer.write_section_title(ws, row, "No Move In", len(headers))
    row, end_row = writer.write_table(ws, row, headers, [build_row(t) for t in without_move_in])
    for i, t in enumerate(without_move_in, start=row - len(without_move_in)):
        _apply_status_fill(writer, ws, i, 3, _status_label(t))
        if t.get("confirmed_move_out_date"):
            writer.apply_fill(ws, i, 7, "green")
    writer.auto_size_columns(ws)

    # 3) Available Units
    ws = writer.new_sheet("Available Units")
    avail_headers = ["Phase", "Unit", "Status", "Available Date", "Move-In Ready Date"]
    available = [t for t in turnovers if _is_vacant(t)]
    writer.write_table(
        ws,
        1,
        avail_headers,
        [[_phase_display(t), _unit_display(t), _status_label(t), parse_date(t.get("available_date")), parse_date(t.get("report_ready_date"))] for t in available],
    )
    for i, t in enumerate(available, start=2):
        _apply_status_fill(writer, ws, i, 3, _status_label(t))
    writer.auto_size_columns(ws)

    # 4) Move Ins
    ws = writer.new_sheet("Move Ins")
    move_ins = [t for t in turnovers if _has_move_in(t)]
    writer.write_table(
        ws,
        1,
        ["Phase", "Unit", "Move In Date"],
        [[_phase_display(t), _unit_display(t), parse_date(t.get("move_in_date"))] for t in move_ins],
    )
    writer.auto_size_columns(ws)

    # 5) Move Outs
    ws = writer.new_sheet("Move Outs")
    writer.write_table(
        ws,
        1,
        ["Phase", "Unit", "Move-Out Date"],
        [[_phase_display(t), _unit_display(t), parse_date(t.get("available_date"))] for t in turnovers],
    )
    writer.auto_size_columns(ws)

    # 6) Pending FAS
    ws = writer.new_sheet("Pending FAS")
    pending = [t for t in turnovers if _is_notice(t)]
    writer.write_table(
        ws,
        1,
        ["Phase", "Unit", "MO/Cancel Date", "Lease End", "Completed"],
        [
            [
                _phase_display(t),
                _unit_display(t),
                parse_date(t.get("available_date")),
                parse_date(t.get("scheduled_move_out_date")),
                "",
            ]
            for t in pending
        ],
    )
    writer.auto_size_columns(ws)

    # 7) Move Activity
    ws = writer.new_sheet("Move Activity")
    writer.write_table(
        ws,
        1,
        ["Phase", "Unit", "Move-Out Date", "Move In Date"],
        [[_phase_display(t), _unit_display(t), parse_date(t.get("available_date")), parse_date(t.get("move_in_date"))] for t in turnovers],
    )
    writer.auto_size_columns(ws)
    return workbook_to_bytes(wb)


def _bucket_label(dv: int) -> str:
    if dv <= 10:
        return "1-10"
    if dv <= 20:
        return "11-20"
    if dv <= 30:
        return "21-30"
    if dv <= 60:
        return "31-60"
    if dv <= 120:
        return "61-120"
    return "120+"


def _write_bucket_grid(writer: WorkbookWriter, ws, row: int, title: str, items: list[dict], *, force_red: bool = False) -> int:
    headers = ["1-10", "11-20", "21-30", "31-60", "61-120", "120+"]
    row = writer.write_section_title(ws, row, title, len(headers))
    row = writer.write_empty_table_header(ws, row, headers)
    grouped: dict[str, list[str]] = {h: [] for h in headers}
    for t in items:
        dv = _safe_int(t.get("dv"), 0)
        label = _bucket_label(dv)
        grouped[label].append(f"Unit {_unit_display(t)} ({_status_label(t)}) | DV-{dv}")
    max_len = max((len(v) for v in grouped.values()), default=0)
    for r_offset in range(max_len):
        for c_idx, h in enumerate(headers, start=1):
            val = grouped[h][r_offset] if r_offset < len(grouped[h]) else ""
            cell = ws.cell(row=row, column=c_idx, value=val)
            if val:
                if force_red:
                    fill_name = "red"
                else:
                    fill_name = "red" if "not ready" in val.lower() else "green"
                writer.apply_fill(ws, row, c_idx, fill_name)
        row += 1
    return row + 1


def _build_dmrb_report(turnovers: list[dict], today: date) -> bytes:
    wb = Workbook()
    writer = WorkbookWriter(wb)

    # 1) Dashboard
    ws = writer.new_sheet("Dashboard")
    dashboard_headers = ["Phase", "Unit", "Status", "Days Vacant", "M-I Date", "Alert"]
    writer.write_table(
        ws,
        1,
        dashboard_headers,
        [
            [_phase_display(t), _unit_display(t), _status_label(t), _safe_int(t.get("dv")), parse_date(t.get("move_in_date")), t.get("attention_badge") or ""]
            for t in turnovers
        ],
    )
    for i, t in enumerate(turnovers, start=2):
        _apply_status_fill(writer, ws, i, 3, _status_label(t))
        writer.apply_fill(ws, i, 4, dv_fill_name(t.get("dv")))
        writer.apply_fill(ws, i, 6, alert_fill_name(t.get("attention_badge") or ""))
    writer.auto_size_columns(ws)

    # 2) Aging
    ws = writer.new_sheet("Aging")
    row = 1
    all_units = [t for t in turnovers if _is_vacant(t) or _safe_int(t.get("dv")) > 0]
    vacant_ready = [t for t in all_units if _status_is_ready(t)]
    vacant_not_ready = [t for t in all_units if _status_is_not_ready(t)]
    row = _write_bucket_grid(writer, ws, row, "Aging Buckets — All Units", all_units)
    row = _write_bucket_grid(writer, ws, row, "Aging Buckets — Vacant Ready", vacant_ready)
    row = _write_bucket_grid(writer, ws, row, "Aging Buckets — Vacant Not Ready", vacant_not_ready, force_red=True)
    for idx in range(1, 7):
        ws.column_dimensions[chr(64 + idx)].width = 45

    # 3) Active Aging
    ws = writer.new_sheet("Active Aging")
    row = 1
    by_phase: dict[str, list[dict]] = defaultdict(list)
    for t in turnovers:
        if _status_is_not_ready(t) and _safe_int(t.get("dv")) > 0:
            by_phase[_phase_display(t) or "Unknown"].append(t)
    for phase in sorted(by_phase.keys()):
        section = by_phase[phase]
        headers = sorted({_bucket_label(_safe_int(t.get("dv"), 0)) for t in section}, key=lambda s: [1, 11, 21, 31, 61, 120].index(int(s.split("-")[0].replace("+", ""))))
        row = writer.write_section_title(ws, row, f"Active Aging — {phase} (Vacant Not Ready)", max(len(headers), 1))
        row = writer.write_empty_table_header(ws, row, headers)
        grouped: dict[str, list[str]] = {h: [] for h in headers}
        for t in section:
            grouped[_bucket_label(_safe_int(t.get("dv"), 0))].append(
                f"Unit {_unit_display(t)} ({_status_label(t)}) | DV-{_safe_int(t.get('dv'), 0)}"
            )
        max_rows = max((len(v) for v in grouped.values()), default=0)
        for _ in range(max_rows):
            for c_idx, h in enumerate(headers, start=1):
                txt = grouped[h].pop(0) if grouped[h] else ""
                ws.cell(row=row, column=c_idx, value=txt)
                if txt:
                    writer.apply_fill(ws, row, c_idx, "red")
            row += 1
        row += 1
    for col in range(1, 25):
        ws.column_dimensions[chr(64 + col if col <= 26 else 64 + (col - 26))].width = 45

    # 4) Operations
    ws = writer.new_sheet("Operations")
    row = 1
    ops_sections = [
        ("Move-In Dashboard", [t for t in turnovers if _has_move_in(t)], ["Phase", "Unit", "Days to Move-In"], lambda t: [_phase_display(t), _unit_display(t), t.get("days_to_move_in")]),
        ("SLA Breach", [t for t in turnovers if bool(t.get("sla_breach"))], ["Phase", "Unit"], lambda t: [_phase_display(t), _unit_display(t)]),
        ("Inspection SLA", [t for t in turnovers if bool(t.get("inspection_sla_breach"))], ["Phase", "Unit"], lambda t: [_phase_display(t), _unit_display(t)]),
        ("Plan Breach", [t for t in turnovers if bool(t.get("plan_breach"))], ["Phase", "Unit"], lambda t: [_phase_display(t), _unit_display(t)]),
        ("Task Stalled", [t for t in turnovers if bool(t.get("is_task_stalled"))], ["Phase", "Unit"], lambda t: [_phase_display(t), _unit_display(t)]),
        ("Clean Turn", [t for t in turnovers if _safe_float(t.get("task_completion_ratio")) >= 100], ["Phase", "Unit"], lambda t: [_phase_display(t), _unit_display(t)]),
    ]
    for title, data, headers, row_builder in ops_sections:
        row = writer.write_section_title(ws, row, title, len(headers))
        if data:
            row, _ = writer.write_table(ws, row, headers, [row_builder(t) for t in data])
        else:
            row = writer.write_empty_table_header(ws, row, headers)
            row = writer.write_empty_message(ws, row, "No data", len(headers))
    writer.auto_size_columns(ws)

    # 5) Walking Path Board
    ws = writer.new_sheet("Walking Path Board")
    wp_headers = ["Phase", "Building", "Unit", "Status", "Move In", "Alert", "Insp", "Paint", "MR", "HK", "CC", "QC", "Notes"]
    writer.write_table(
        ws,
        1,
        wp_headers,
        [
            [
                _phase_display(t),
                _phase_display(t),
                _unit_display(t),
                _status_label(t),
                parse_date(t.get("move_in_date")),
                t.get("attention_badge") or "",
                _task_execution_status(t, "task_insp"),
                _task_execution_status(t, "task_paint"),
                _task_execution_status(t, "task_mr"),
                _task_execution_status(t, "task_hk"),
                _task_execution_status(t, "task_cc"),
                t.get("qc_status") or "",
                t.get("notes_joined") or "",
            ]
            for t in turnovers
        ],
    )
    for i, t in enumerate(turnovers, start=2):
        _apply_status_fill(writer, ws, i, 4, _status_label(t))
        writer.apply_fill(ws, i, 6, alert_fill_name(t.get("attention_badge") or ""))
        writer.apply_fill(ws, i, 7, task_status_fill_name(_task_execution_status(t, "task_insp")))
        writer.apply_fill(ws, i, 8, task_status_fill_name(_task_execution_status(t, "task_paint")))
        writer.apply_fill(ws, i, 9, task_status_fill_name(_task_execution_status(t, "task_mr")))
        writer.apply_fill(ws, i, 10, task_status_fill_name(_task_execution_status(t, "task_hk")))
        writer.apply_fill(ws, i, 11, task_status_fill_name(_task_execution_status(t, "task_cc")))
    writer.auto_size_columns(ws)

    # 6) Tasks
    ws = writer.new_sheet("Tasks")
    rows = []
    for t in turnovers:
        current = str(t.get("current_task") or "")
        next_task = str(t.get("next_task") or "")
        current_key = _task_key_from_label(current)
        next_key = _task_key_from_label(next_task)
        rows.append(
            [
                _phase_display(t),
                _unit_display(t),
                current,
                _task_due_date(t, current_key) if current_key else None,
                next_task,
                _task_due_date(t, next_key) if next_key else None,
                _safe_float(t.get("task_completion_ratio")),
            ]
        )
    writer.write_table(ws, 1, ["Phase", "Unit", "Current Task", "Task Date", "Next", "Next Date", "Progress %"], rows)
    for i, t in enumerate(turnovers, start=2):
        writer.apply_fill(ws, i, 7, progress_fill_name(t.get("task_completion_ratio")))
    writer.auto_size_columns(ws)

    # 7) Schedule
    ws = writer.new_sheet("Schedule")
    row = 1
    for label, key in TASK_COLUMNS:
        row = writer.write_section_title(ws, row, label, 3)
        filtered = [t for t in turnovers if parse_date((t.get(key) or {}).get("scheduled_date"))]
        if filtered:
            row, _ = writer.write_table(
                ws,
                row,
                ["Phase", "Unit", "Date"],
                [[_phase_display(t), _unit_display(t), parse_date((t.get(key) or {}).get("scheduled_date"))] for t in filtered],
            )
        else:
            row = writer.write_empty_table_header(ws, row, ["Phase", "Unit", "Date"])
            row = writer.write_empty_message(ws, row, "No scheduled items", 3)
    writer.auto_size_columns(ws)

    # 8) Upcoming
    ws = writer.new_sheet("Upcoming")
    row = 1
    for label, key in TASK_COLUMNS:
        row = writer.write_section_title(ws, row, label, 3)
        filtered = []
        for t in turnovers:
            due = parse_date((t.get(key) or {}).get("vendor_due_date"))
            if due is None:
                continue
            delta = (due - today).days
            if 0 <= delta <= 7:
                filtered.append(t)
        if filtered:
            row, _ = writer.write_table(
                ws,
                row,
                ["Phase", "Unit", "Date"],
                [[_phase_display(t), _unit_display(t), parse_date((t.get(key) or {}).get("vendor_due_date"))] for t in filtered],
            )
        else:
            row = writer.write_empty_table_header(ws, row, ["Phase", "Unit", "Date"])
            row = writer.write_empty_message(ws, row, "No upcoming items", 3)
    writer.auto_size_columns(ws)

    # 9) WD Audit
    ws = writer.new_sheet("WD Audit")
    writer.write_table(
        ws,
        1,
        ["Phase", "Unit", "W/D"],
        [[_phase_display(t), _unit_display(t), t.get("wd_summary") or ""] for t in turnovers],
    )
    for i, t in enumerate(turnovers, start=2):
        writer.apply_fill(ws, i, 3, wd_fill_name(t.get("wd_summary") or ""))
    writer.auto_size_columns(ws)

    # 10) Daily Ops
    ws = writer.new_sheet("Daily Ops")
    row = 1
    row = writer.write_section_title(ws, row, "Portfolio Overview", 2)
    overview = [
        ("Total Units", len(turnovers)),
        ("Vacant", sum(1 for t in turnovers if _is_vacant(t))),
        ("On Notice", sum(1 for t in turnovers if _is_notice(t))),
    ]
    row, _ = writer.write_table(ws, row, ["Metric", "Value"], overview, as_table=False)
    row += 1
    row = writer.write_section_title(ws, row, "Turn Performance", 2)
    avg_dv = mean([_safe_float(t.get("dv")) for t in turnovers]) if turnovers else 0.0
    avg_completion = mean([_safe_float(t.get("task_completion_ratio")) for t in turnovers]) if turnovers else 0.0
    perf = [
        ("Avg Days Vacant", round(avg_dv, 2)),
        ("Avg Completion %", round(avg_completion, 2)),
        ("SLA Breaches", sum(1 for t in turnovers if t.get("sla_breach"))),
        ("Plan Breaches", sum(1 for t in turnovers if t.get("plan_breach"))),
        ("Ready Units", sum(1 for t in turnovers if t.get("is_unit_ready"))),
    ]
    writer.write_table(ws, row, ["Metric", "Value"], perf, as_table=False)
    writer.auto_size_columns(ws)

    # 11) Priority
    ws = writer.new_sheet("Priority")
    priority_rows = []
    for t in turnovers:
        reasons = []
        if t.get("operational_state") == "Move-In Risk":
            reasons.append("Move-In Risk")
        if t.get("sla_breach"):
            reasons.append("SLA Breach")
        if t.get("plan_breach"):
            reasons.append("Plan Breach")
        priority_rows.append([_phase_display(t), _unit_display(t), t.get("attention_badge") or "", " / ".join(reasons)])
    writer.write_table(ws, 1, ["Phase", "Unit", "Priority_Flag", "Urgency_Reason"], priority_rows)
    for i, t in enumerate(turnovers, start=2):
        writer.apply_fill(ws, i, 3, alert_fill_name(t.get("attention_badge") or ""))
        if ws.cell(row=i, column=4).value:
            writer.apply_fill(ws, i, 4, "yellow")
    writer.auto_size_columns(ws)

    # 12) Phase Performance
    ws = writer.new_sheet("Phase Performance")
    by_phase_perf: dict[str, list[dict]] = defaultdict(list)
    for t in turnovers:
        by_phase_perf[_phase_display(t) or "Unknown"].append(t)
    perf_rows = []
    for phase in sorted(by_phase_perf.keys()):
        group = by_phase_perf[phase]
        count = len(group)
        avg_dv = mean(_safe_float(t.get("dv")) for t in group) if group else 0.0
        completion = mean(_safe_float(t.get("task_completion_ratio")) for t in group) if group else 0.0
        non_breach = sum(1 for t in group if not t.get("sla_breach"))
        sla_pct = (non_breach / count) * 100 if count else 0.0
        perf_rows.append([phase, count, round(avg_dv, 2), round(sla_pct, 2), round(completion, 2)])
    writer.write_table(ws, 1, ["Phase", "Count", "Avg DV", "SLA Compliance %", "Completion Rate"], perf_rows)
    for i, row in enumerate(perf_rows, start=2):
        writer.apply_fill(ws, i, 4, sla_compliance_fill_name(row[3]))
    writer.auto_size_columns(ws)

    return workbook_to_bytes(wb)


def _build_dashboard_chart(turnovers: list[dict]) -> bytes:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, axes = plt.subplots(3, 3, figsize=(16, 12), dpi=100)
    ax = axes.ravel()

    def _plot_bar(axis, title: str, labels: list[str], values: list[float]):
        axis.bar(labels, values)
        axis.set_title(title)
        axis.tick_params(axis="x", labelrotation=30)

    # 1 Turn Time Distribution
    turn_buckets = {"0-5": 0, "6-10": 0, "11-15": 0, "16-20": 0, "21+": 0}
    for t in turnovers:
        dv = _safe_int(t.get("dv"), 0)
        if dv <= 5:
            turn_buckets["0-5"] += 1
        elif dv <= 10:
            turn_buckets["6-10"] += 1
        elif dv <= 15:
            turn_buckets["11-15"] += 1
        elif dv <= 20:
            turn_buckets["16-20"] += 1
        else:
            turn_buckets["21+"] += 1
    _plot_bar(ax[0], "Turn Time Distribution", list(turn_buckets.keys()), list(turn_buckets.values()))

    # 2 Units by State
    states: dict[str, int] = defaultdict(int)
    for t in turnovers:
        states[str(t.get("operational_state") or "Unknown")] += 1
    _plot_bar(ax[1], "Units by State", list(states.keys()), list(states.values()))

    # 3 Task Completion
    comp = {"0-25%": 0, "26-50%": 0, "51-75%": 0, "76-100%": 0}
    for t in turnovers:
        p = _safe_float(t.get("task_completion_ratio"))
        if p <= 25:
            comp["0-25%"] += 1
        elif p <= 50:
            comp["26-50%"] += 1
        elif p <= 75:
            comp["51-75%"] += 1
        else:
            comp["76-100%"] += 1
    _plot_bar(ax[2], "Task Completion", list(comp.keys()), list(comp.values()))

    by_phase: dict[str, list[dict]] = defaultdict(list)
    for t in turnovers:
        by_phase[_phase_display(t) or "Unknown"].append(t)
    phases = sorted(by_phase.keys())

    # 4 SLA by Phase
    _plot_bar(ax[3], "SLA by Phase", phases, [sum(1 for t in by_phase[p] if t.get("sla_breach")) for p in phases])
    # 5 Avg Turn Time by Phase
    _plot_bar(ax[4], "Avg Turn Time by Phase", phases, [mean(_safe_float(t.get("dv")) for t in by_phase[p]) if by_phase[p] else 0 for p in phases])
    # 6 Task Completion Rate
    _plot_bar(ax[5], "Task Completion Rate", phases, [mean(_safe_float(t.get("task_completion_ratio")) for t in by_phase[p]) if by_phase[p] else 0 for p in phases])
    # 7 Vacancy Rate by Phase
    _plot_bar(ax[6], "Vacancy Rate by Phase", phases, [sum(1 for t in by_phase[p] if _is_vacant(t)) for p in phases])

    # 8 Units by Badge
    badges: dict[str, int] = defaultdict(int)
    for t in turnovers:
        badges[str(t.get("attention_badge") or "Unknown")] += 1
    _plot_bar(ax[7], "Units by Badge", list(badges.keys()), list(badges.values()))

    # 9 Days Until Move-In
    movein_buckets = {"0-2": 0, "3-7": 0, "8-14": 0, "15+": 0}
    for t in turnovers:
        d = t.get("days_to_move_in")
        if d is None:
            continue
        d = _safe_int(d)
        if d <= 2:
            movein_buckets["0-2"] += 1
        elif d <= 7:
            movein_buckets["3-7"] += 1
        elif d <= 14:
            movein_buckets["8-14"] += 1
        else:
            movein_buckets["15+"] += 1
    _plot_bar(ax[8], "Days Until Move-In", list(movein_buckets.keys()), list(movein_buckets.values()))

    fig.tight_layout(pad=3.0)
    buf = BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    return buf.getvalue()


def _build_weekly_summary(turnovers: list[dict], today: date) -> str:
    total = len(turnovers)
    vacant = sum(1 for t in turnovers if _is_vacant(t))
    on_notice = sum(1 for t in turnovers if _is_notice(t))
    sla_breaches = [t for t in turnovers if t.get("sla_breach")]
    plan_breaches = [t for t in turnovers if t.get("plan_breach")]
    ready_units = sum(1 for t in turnovers if t.get("is_unit_ready"))
    avg_dv = round(mean([_safe_float(t.get("dv")) for t in turnovers]), 2) if turnovers else 0.0
    sla_compliance = round(((total - len(sla_breaches)) / total) * 100, 2) if total else 100.0

    move_in_risk = [t for t in turnovers if str(t.get("operational_state") or "") == "Move-In Risk"]
    stalled = [t for t in turnovers if bool(t.get("is_task_stalled"))]
    upcoming_moveins = [t for t in turnovers if t.get("days_to_move_in") is not None and 0 <= _safe_int(t.get("days_to_move_in")) <= 7]
    upcoming_ready = []
    for t in turnovers:
        ready_date = parse_date(t.get("report_ready_date"))
        if ready_date is None:
            continue
        delta = (ready_date - today).days
        if 0 <= delta <= 7:
            upcoming_ready.append(t)
    wd_not_installed = [t for t in turnovers if bool(t.get("wd_present")) and not bool(t.get("wd_installed"))]

    aging = {"1-10": 0, "11-20": 0, "21-30": 0, "31-60": 0, "61-120": 0, "120+": 0}
    for t in turnovers:
        dv = _safe_int(t.get("dv"), 0)
        aging[_bucket_label(dv)] += 1

    def units_list(items: list[dict], *, include_due: bool = False, include_task: bool = False) -> str:
        if not items:
            return "None"
        lines = []
        for t in items:
            unit = _unit_display(t)
            if include_due:
                due = parse_date(t.get("move_in_date"))
                lines.append(f"- {unit} (due {due.isoformat() if due else 'N/A'})")
            elif include_task:
                lines.append(f"- {unit} (current task: {t.get('current_task') or 'N/A'})")
            else:
                lines.append(f"- {unit}")
        return "\n".join(lines)

    return (
        "KEY METRICS\n"
        f"Total Active: {total}\n"
        f"Vacant: {vacant}\n"
        f"On Notice: {on_notice}\n"
        f"SLA Compliance %: {sla_compliance}\n"
        f"Ready Units: {ready_units}\n"
        f"Avg Days Vacant: {avg_dv}\n\n"
        "ALERTS\n"
        f"SLA Breaches:\n{units_list(sla_breaches)}\n"
        f"Move-In Risk:\n{units_list(move_in_risk, include_due=True)}\n"
        f"Stalled Tasks:\n{units_list(stalled, include_task=True)}\n"
        f"Plan Breaches:\n{units_list(plan_breaches)}\n\n"
        "UPCOMING MOVE-INS\n"
        f"{units_list(upcoming_moveins, include_due=True)}\n\n"
        "UPCOMING READY\n"
        f"{units_list(upcoming_ready)}\n\n"
        "W/D NOT INSTALLED\n"
        f"{units_list(wd_not_installed)}\n\n"
        "AGING DISTRIBUTION\n"
        f"1-10: {aging['1-10']}\n"
        f"11-20: {aging['11-20']}\n"
        f"21-30: {aging['21-30']}\n"
        f"31-60: {aging['31-60']}\n"
        f"61-120: {aging['61-120']}\n"
        f"120+: {aging['120+']}\n"
    )


def generate_final_report(turnovers: list[dict]) -> bytes:
    return _build_final_report(turnovers)


def generate_dmrb_report(turnovers: list[dict], today: date | None = None) -> bytes:
    return _build_dmrb_report(turnovers, today or date.today())


def generate_dashboard_chart(turnovers: list[dict]) -> bytes:
    return _build_dashboard_chart(turnovers)


def generate_weekly_summary(turnovers: list[dict], today: date | None = None) -> bytes:
    txt = _build_weekly_summary(turnovers, today or date.today())
    return txt.encode("utf-8")


def generate_all_reports_zip(turnovers: list[dict], today: date | None = None) -> bytes:
    today = today or date.today()
    with ThreadPoolExecutor(max_workers=4) as executor:
        fut_final = executor.submit(generate_final_report, turnovers)
        fut_dmrb = executor.submit(generate_dmrb_report, turnovers, today)
        fut_png = executor.submit(generate_dashboard_chart, turnovers)
        fut_txt = executor.submit(generate_weekly_summary, turnovers, today)
        final_bytes = fut_final.result()
        dmrb_bytes = fut_dmrb.result()
        png_bytes = fut_png.result()
        txt_bytes = fut_txt.result()

    buf = BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        zf.writestr(FINAL_REPORT_FILENAME, final_bytes)
        zf.writestr(DMRB_REPORT_FILENAME, dmrb_bytes)
        zf.writestr(DASHBOARD_CHART_FILENAME, png_bytes)
        zf.writestr(WEEKLY_SUMMARY_FILENAME, txt_bytes)
    return buf.getvalue()


def generate_all_export_artifacts(conn, today: date | None = None) -> dict[str, bytes]:
    today = today or date.today()
    turnovers = build_export_turnovers(conn, today=today)
    final_bytes = generate_final_report(turnovers)
    dmrb_bytes = generate_dmrb_report(turnovers, today=today)
    chart_bytes = generate_dashboard_chart(turnovers)
    weekly_bytes = generate_weekly_summary(turnovers, today=today)
    all_zip = generate_all_reports_zip(turnovers, today=today)
    return {
        FINAL_REPORT_FILENAME: final_bytes,
        DMRB_REPORT_FILENAME: dmrb_bytes,
        DASHBOARD_CHART_FILENAME: chart_bytes,
        WEEKLY_SUMMARY_FILENAME: weekly_bytes,
        ALL_REPORTS_FILENAME: all_zip,
    }
