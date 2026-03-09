from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import zipfile

from openpyxl import load_workbook

from db import repository
from services import export_service
from services.excel_writer import (
    alert_fill_name,
    dv_fill_name,
    progress_fill_name,
    sla_compliance_fill_name,
    status_fill_name,
    task_status_fill_name,
    wd_fill_name,
)
from tests.helpers.db_bootstrap import runtime_db


def _seed_dataset(conn):
    now = datetime.utcnow().isoformat()
    conn.execute("INSERT INTO property (property_id, name) VALUES (?, ?)", (1, "P1"))
    unit_1 = repository.resolve_unit(
        conn,
        property_id=1,
        phase_code="5",
        building_code="18",
        unit_number="0101",
        unit_code_raw="5-18-0101",
        unit_code_norm="5-18-0101",
        unit_identity_key="5-18-0101",
    )
    unit_2 = repository.resolve_unit(
        conn,
        property_id=1,
        phase_code="7",
        building_code="22",
        unit_number="0202",
        unit_code_raw="7-22-0202",
        unit_code_norm="7-22-0202",
        unit_identity_key="7-22-0202",
    )
    unit_3 = repository.resolve_unit(
        conn,
        property_id=1,
        phase_code="8",
        building_code="30",
        unit_number="0303",
        unit_code_raw="8-30-0303",
        unit_code_norm="8-30-0303",
        unit_identity_key="8-30-0303",
    )

    t1 = repository.insert_turnover(
        conn,
        {
            "property_id": 1,
            "unit_id": unit_1["unit_id"],
            "source_turnover_key": "T1",
            "move_out_date": "2026-03-01",
            "move_in_date": "2026-03-20",
            "report_ready_date": "2026-03-15",
            "manual_ready_status": "Vacant not ready",
            "wd_present": 1,
            "wd_supervisor_notified": 1,
            "wd_installed": 0,
            "created_at": now,
            "updated_at": now,
            "confirmed_move_out_date": "2026-03-02",
            "scheduled_move_out_date": "2026-03-03",
        },
    )
    t2 = repository.insert_turnover(
        conn,
        {
            "property_id": 1,
            "unit_id": unit_2["unit_id"],
            "source_turnover_key": "T2",
            "move_out_date": "2026-03-04",
            "move_in_date": None,
            "report_ready_date": "2026-03-25",
            "manual_ready_status": "On notice",
            "wd_present": 0,
            "wd_supervisor_notified": 0,
            "wd_installed": 0,
            "created_at": now,
            "updated_at": now,
            "scheduled_move_out_date": "2026-03-30",
        },
    )
    repository.insert_turnover(
        conn,
        {
            "property_id": 1,
            "unit_id": unit_3["unit_id"],
            "source_turnover_key": "T3",
            "move_out_date": "2026-03-01",
            "move_in_date": None,
            "report_ready_date": None,
            "manual_ready_status": "Vacant not ready",
            "wd_present": 1,
            "wd_supervisor_notified": 0,
            "wd_installed": 0,
            "created_at": now,
            "updated_at": now,
            "closed_at": now,
        },
    )

    for tid in (t1, t2):
        for task_type, exec_status in (
            ("Insp", "VENDOR_COMPLETED"),
            ("Paint", "IN_PROGRESS"),
            ("MR", "NOT_STARTED"),
            ("HK", "NOT_STARTED"),
            ("CC", "NOT_STARTED"),
            ("QC", "NOT_STARTED"),
        ):
            repository.insert_task(
                conn,
                {
                    "turnover_id": tid,
                    "task_type": task_type,
                    "required": 1,
                    "blocking": 1,
                    "execution_status": exec_status,
                    "confirmation_status": "PENDING",
                    "vendor_completed_at": now if exec_status == "VENDOR_COMPLETED" else None,
                },
            )
    repository.insert_task(
        conn,
        {
            "turnover_id": t1,
            "task_type": "FW",
            "required": 1,
            "blocking": 1,
            "execution_status": "VENDOR_COMPLETED",
            "confirmation_status": "CONFIRMED",
            "vendor_completed_at": now,
            "manager_confirmed_at": now,
        },
    )
    repository.insert_note(
        conn,
        {
            "turnover_id": t1,
            "note_type": "INFO",
            "blocking": 0,
            "severity": "INFO",
            "description": "Waiting on key delivery",
            "created_at": now,
        },
    )
    repository.insert_note(
        conn,
        {
            "turnover_id": t2,
            "note_type": "INFO",
            "blocking": 0,
            "severity": "INFO",
            "description": "Pending lease docs",
            "created_at": now,
        },
    )
    conn.commit()


def test_build_export_turnovers_derives_fields_and_scope():
    with runtime_db() as (conn, _):
        _seed_dataset(conn)
        rows = export_service.build_export_turnovers(conn, today=date(2026, 3, 10))
        assert len(rows) == 2  # closed turnover excluded

        by_unit = {r["unit_code"]: r for r in rows}
        a = by_unit["5-18-0101"]
        b = by_unit["7-22-0202"]

        assert a["qc_status"] == "Confirmed"
        assert b["qc_status"] == "N/A"
        assert a["wd_summary"] == "NOTIFIED"
        assert b["wd_summary"] == "—"
        assert a["notes_joined"] == "Waiting on key delivery"


def test_generate_final_report_workbook_shape():
    with runtime_db() as (conn, _):
        _seed_dataset(conn)
        rows = export_service.build_export_turnovers(conn, today=date(2026, 3, 10))
        payload = export_service.generate_final_report(rows)
        wb = load_workbook(BytesIO(payload))
        assert wb.sheetnames == [
            "Reconciliation",
            "Split View",
            "Available Units",
            "Move Ins",
            "Move Outs",
            "Pending FAS",
            "Move Activity",
        ]
        ws = wb["Reconciliation"]
        assert ws["A1"].value == "Phase"
        assert ws["B1"].value == "Unit"
        assert ws["G1"].value == "MO/Confirm"


def test_generate_dmrb_report_workbook_shape():
    with runtime_db() as (conn, _):
        _seed_dataset(conn)
        rows = export_service.build_export_turnovers(conn, today=date(2026, 3, 10))
        payload = export_service.generate_dmrb_report(rows, today=date(2026, 3, 10))
        wb = load_workbook(BytesIO(payload))
        assert wb.sheetnames == [
            "Dashboard",
            "Aging",
            "Active Aging",
            "Operations",
            "Walking Path Board",
            "Tasks",
            "Schedule",
            "Upcoming",
            "WD Audit",
            "Daily Ops",
            "Priority",
            "Phase Performance",
        ]
        ws = wb["Dashboard"]
        assert ws["A1"].value == "Phase"
        assert ws["D1"].value == "Days Vacant"


def test_generate_png_txt_and_zip_outputs():
    with runtime_db() as (conn, _):
        _seed_dataset(conn)
        rows = export_service.build_export_turnovers(conn, today=date(2026, 3, 10))
        png_payload = export_service.generate_dashboard_chart(rows)
        assert png_payload[:8] == b"\x89PNG\r\n\x1a\n"

        txt_payload = export_service.generate_weekly_summary(rows, today=date(2026, 3, 10))
        txt = txt_payload.decode("utf-8")
        assert "KEY METRICS" in txt
        assert "ALERTS" in txt
        assert "AGING DISTRIBUTION" in txt

        zip_payload = export_service.generate_all_reports_zip(rows, today=date(2026, 3, 10))
        with zipfile.ZipFile(BytesIO(zip_payload), "r") as zf:
            names = sorted(zf.namelist())
        assert names == sorted(
            [
                "Final_Report.xlsx",
                "DMRB_Report.xlsx",
                "Dashboard_Chart.png",
                "Weekly_Summary.txt",
            ]
        )


def test_fill_mapping_helpers():
    assert status_fill_name("Vacant ready") == "green"
    assert status_fill_name("Vacant not ready") == "red"
    assert status_fill_name("On notice") == "gray"

    assert alert_fill_name("CRITICAL MOVE-IN RISK") == "red"
    assert alert_fill_name("In Progress") == "blue"
    assert alert_fill_name("On Notice") == "gray"

    assert dv_fill_name(3) == "green"
    assert dv_fill_name(10) == "amber"
    assert dv_fill_name(25) == "red"

    assert progress_fill_name(90) == "green"
    assert progress_fill_name(60) == "amber"
    assert progress_fill_name(10) == "red"

    assert task_status_fill_name("VENDOR_COMPLETED") == "green"
    assert task_status_fill_name("IN_PROGRESS") == "blue"
    assert task_status_fill_name("NOT_STARTED") == "red"

    assert wd_fill_name("OK") == "green"
    assert wd_fill_name("NOTIFIED") == "amber"
    assert wd_fill_name("PENDING") == "red"

    assert sla_compliance_fill_name(95) == "green"
    assert sla_compliance_fill_name(80) == "yellow"
    assert sla_compliance_fill_name(60) == "red"
